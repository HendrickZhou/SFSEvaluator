"""
All the functions need to calculate the metrics

@Author: Hang Zhou
"""

from cg_util import ThreeDObject
from openpyxl import Workbook,load_workbook
from openpyxl.drawing.image import Image
from sfseval.util import file_exist,create_dir
from config import METRIC_PATH
import numpy as np
from sfseval.util import *


class MetricDataObject:
    def __init__(self) -> None:
        self.tags=set()
        self.algorithm=None
        self.dataset=None
        self.MAE=None
        self.RMSE=None
        self.depth_image=None # must be a legit hyperlink
        self.gt_image=None # must be a legit hyperlink

    def __eq__(self, other) -> bool:
        if self.algorithm!=other.algorithm:
            return False
        if self.dataset!=other.dataset:
            return False
        if len(self.tags) != len(other.tags):
            return False
        if self.tags != other.tags:
            return False
        return True
        
    
class MetricsDB:
    xls_name=METRIC_PATH+"metric.xlsx"
    headings=["algorithm","dataset","tags", "depth_image","gt_image","MAE","RMSE"]
    dm_idx=3
    gm_idx=4
    def __init__(self) -> None:
        self.wb=None
        create_dir(abs_name=METRIC_PATH)

    def __enter__(self):
        print("connect to metric data")
        self._start_wb()
        return self

    def __exit__(self,type, value, traceback):
        print("saving new metrics data............")
        print("close connection to metrics data")
        self.wb.save(self.xls_name)

    def _start_wb(self):
        if file_exist(abs_name=self.xls_name):
            self.wb=load_workbook(filename=self.xls_name)
        else:
            wb=Workbook()
            ws=wb.active
            ws.append(self.headings)
            self.wb=wb

    def _m2list(self, m:MetricDataObject):
        result=[None for _ in range(len(self.headings))]
        result[0]=m.algorithm
        result[1]=m.dataset
        tag_str=""
        for tag in m.tags:
            tag_str+=tag
        result[2]=tag_str
        result[3]=None
        result[4]=None
        result[5]=m.MAE
        result[6]=m.RMSE
        return result

    def add_result(self,m:MetricDataObject):
        worksheet=self.wb.active
        worksheet.append(self._m2list(m))
        last_row=len(list(worksheet.rows))
        # handle image
        if m.depth_image is not None:
            the_cell=worksheet.cell(last_row,self.dm_idx+1)
            the_cell.hyperlink=m.depth_image
            the_cell.value="link_depth"
            the_cell.style="Hyperlink"
        if m.gt_image is not None:
            the_cell=worksheet.cell(last_row,self.gm_idx+1)
            the_cell.hyperlink=m.gt_image
            the_cell.value="link_gt"
            the_cell.style="Hyperlink"

# cal all metrics
def vis_metrics(output:ThreeDObject,algorithm,dataset,tags):
    tag_str=""
    for tag in tags:
        tag_str+=str(tag)
    img_path=algorithm+"_"+dataset+"_"+tag_str+"/"
    img_path=create_dir(ondir=METRIC_PATH, name=img_path)
    depth_path=img_path+"output_depth.png" 
    output.vis_and_save_depth(depth_path) 
    m=MetricDataObject()
    m.algorithm=algorithm
    m.dataset=dataset
    m.tags=tags
    m.depth_image=depth_path
    return m 

def get_metrics(output:ThreeDObject, ground_truth:ThreeDObject, algorithm, dataset, tags):
    """
    algorithm should be string, and unique
    dataset should be string and unique
    tags should be set of string!
    """
    mae=MAE(output_normal=output.get_normal(),
        gt_normal=ground_truth.get_normal())
    
    rmse=RMSE(output_dm=output.get_depth(),
         gt_dm=ground_truth.get_depth())
    print("MAE="+str(mae))
    print("RMSE="+str(rmse))
    # save the images
    tag_str=""
    for tag in tags:
        tag_str+=str(tag)
    img_path=algorithm+"_"+dataset+"_"+tag_str+"/"
    img_path=create_dir(ondir=METRIC_PATH, name=img_path)
    depth_path=img_path+"output_depth.png"
    gt_path=img_path+"gt_depth.png"
    output.vis_and_save_depth(depth_path)
    ground_truth.vis_and_save_depth(gt_path)

    m=MetricDataObject()
    m.algorithm=algorithm
    m.dataset=dataset
    m.tags=tags
    m.MAE=mae
    m.RMSE=rmse
    m.depth_image=depth_path
    m.gt_image=gt_path
    return m

def MAE(output_normal:np.ma,gt_normal:np.ma)->float:
    """
    mean angular error
    assume those images are all normalized

    calculated by this formula:
    e = atan2(norm(cross(el,ee)),dot(el,ee));
    """
    output_norm=output_normal.compressed()
    size=output_norm.size
    sum=0
    (w,h,_)=output_normal.shape
    for i in range(w):
        for j in range(h):
            output_norm_ij=output_normal[i,j].data
            gt_norm_ij=gt_normal[i,j].data
            e=np.arctan2(np.linalg.norm(np.cross(output_norm_ij,gt_norm_ij)),
                     np.dot(output_norm_ij,gt_norm_ij)) 
            sum+=e
    return sum/size

def RMSE(output_dm:np.ma, gt_dm:np.ma)->float:
    """
    root mean square error
    assume those images are all normalized
    """
    output_d=output_dm.compressed()
    gt_d=gt_dm.compressed()
    try:
        assert output_d.size == gt_d.size
    except AssertionError as ae:
        raise
    finally:
        print("gt_d.size="+str(gt_d.size))
        print("output_d.size="+str(output_d.size)) 
    assert output_d.size == gt_d.size
    return np.sqrt(np.mean((output_d-gt_d)**2))

if __name__ == "__main__":
    import pickle
    m=pickle.load(open("test_m.p","rb"))
    import pdb; pdb.set_trace()
    with MetricsDB() as mdb:
        mdb.add_result(m)
